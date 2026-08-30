from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas import DataFrame, ExcelWriter

# ACCOUNTING_NUMBER_FORMAT_EUR = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'
# ACCOUNTING_NUMBER_FORMAT_EUR = '#,##0.00 €;-[Red]-#,##0.00 €;0.00 €'
ACCOUNTING_NUMBER_FORMAT_EUR = '_-* #,##0.00 €_-;_-* (#,##0.00) €_-;_-* 0.00 €_-;_-@ _-'
PERCENTAGE_NUMBER_FORMAT = "0.00 %"


def to_excel_table(
        df: DataFrame, 
        excel_path: Path, 
        sheet_name: str = 'Sheet1', 
        table_name: str = 'Table1', 
        index: bool = False,
        col_width: int | None = None,
        accnt_frmt_cols: list[str] | None = None,
        prcnt_frmt_cols: list[str] | None = None,

    ) -> None:
    
    df.to_excel(excel_path, sheet_name=sheet_name, index=index)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    # Bereich der Tabelle bestimmen (inkl. Header‑Zeile)
    first_col = 1
    first_row = 1                     # Header liegt in Zeile 1
    last_row  = ws.max_row            # letzte Zeile mit Daten
    last_col  = ws.max_column         # letzte Spalte

    table_ref = f"{ws.cell(row=first_row, column=first_col).coordinate}:{ws.cell(row=last_row, column=last_col).coordinate}"
    excel_table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    excel_table.tableStyleInfo = style

    ws.add_table(excel_table)
    if col_width is not None:
        for col in range(1, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = col_width

    if accnt_frmt_cols or prcnt_frmt_cols:
        # Mapping: Spaltenname -> Spaltenindex im Worksheet
        header_map = {
            ws.cell(row=first_row, column=c).value: c
            for c in range(first_col, last_col + 1)
        }

        fmt_map: dict[int, str] = {}
        for name in (accnt_frmt_cols or []):
            if name in header_map:
                fmt_map[header_map[name]] = ACCOUNTING_NUMBER_FORMAT_EUR
        for name in (prcnt_frmt_cols or []):
            if name in header_map:
                fmt_map[header_map[name]] = PERCENTAGE_NUMBER_FORMAT

        for col_idx, fmt in fmt_map.items():
            for row in range(first_row + 1, last_row + 1):  # Header überspringen
                ws.cell(row=row, column=col_idx).number_format = fmt    

    wb.save(excel_path)



@dataclass
class SheetConfiguration:
    """
    Configuration for a single worksheet within the Excel file.

    Attributes
    ----------
    df : DataFrame
        The DataFrame to write to this worksheet.
    sheet_name : str
        Name of the worksheet in the Excel file.
    accnt_frmt_cols : list[str]
        Column names to format with accounting number style (EUR currency).
    prcnt_frmt_cols : list[str]
        Column names to format as percentages (values as fractions).
    """

    df: DataFrame
    sheet_name: str
    accnt_frmt_cols: list[str] | None = None
    prcnt_frmt_cols: list[str] | None = None
    col_width: int | None = None


def _column_index_to_excel_letters(column_index: int) -> str:
    """Convert a 1-based column index to its Excel letter representation."""
    letters = ""
    while column_index > 0:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def write_multiple_sheets_to_excel(
    sheet_configurations: list[SheetConfiguration],
    output_file_path: str,
    table_style_name: str = "TableStyleMedium9",
) -> None:
    """
    Write multiple DataFrames to separate worksheets in a single Excel file.
    Each worksheet becomes a styled Excel Table with optional per-column
    number formatting.

    Parameters
    ----------
    sheet_configurations : list[SheetConfiguration]
        List of configuration objects, one per worksheet. Each contains the
        DataFrame, sheet name, and optional formatting column lists.
    output_file_path : str
        Path to the destination .xlsx file.
    table_style_name : str
        OpenPyXL table style name (default: "TableStyleMedium2").
    """

    # Collect all configured column names for validation
    all_accounting_columns: set[str] = set()
    all_percentage_columns: set[str] = set()

    for configuration in sheet_configurations:
        if configuration.accnt_frmt_cols:
            all_accounting_columns.update(configuration.accnt_frmt_cols)
        if configuration.prcnt_frmt_cols:
            all_percentage_columns.update(configuration.prcnt_frmt_cols)

    # Validate that specified columns exist in their respective DataFrames
    for configuration in sheet_configurations:
        data_frame = configuration.df
        sheet_name = configuration.sheet_name

        for column_name in all_accounting_columns:
            if (
                column_name in (configuration.accnt_frmt_cols or [])
                and column_name not in data_frame.columns
            ):
                raise ValueError(
                    f"Accounting column '{column_name}' not found in sheet "
                    f"'{sheet_name}'."
                )

        for column_name in all_percentage_columns:
            if (
                column_name in (configuration.prcnt_frmt_cols or [])
                and column_name not in data_frame.columns
            ):
                raise ValueError(
                    f"Percentage column '{column_name}' not found in sheet "
                    f"'{sheet_name}'."
                )


    with ExcelWriter(output_file_path, engine="openpyxl") as excel_writer:

        for configuration in sheet_configurations:
            current_data_frame = configuration.df
            current_sheet_name = configuration.sheet_name
            current_accounting_columns = configuration.accnt_frmt_cols or []
            current_percentage_columns = configuration.prcnt_frmt_cols or []
            current_col_width = configuration.col_width

            # Write DataFrame to the worksheet
            current_data_frame.to_excel(
                excel_writer, sheet_name=current_sheet_name, index=False
            )

            # Access the worksheet object for further formatting
            current_worksheet = excel_writer.sheets[current_sheet_name]

            # Calculate dimensions of the data range
            number_of_rows = len(current_data_frame)
            number_of_columns = len(current_data_frame.columns)
            end_column_letter = _column_index_to_excel_letters(number_of_columns)
            data_range_reference = (f"A1:{end_column_letter}{number_of_rows + 1}")

            # Create and add an Excel Table with styling
            table_display_name = (current_sheet_name.replace(" ", "_") + "Table")
            excel_table = Table(displayName=table_display_name, ref=data_range_reference)
            excel_table.tableStyleInfo = TableStyleInfo(
                name=table_style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            current_worksheet.add_table(excel_table)

            for column_index, column_name in enumerate(current_data_frame.columns, start=1):
                if column_name in current_accounting_columns:
                    target_format = ACCOUNTING_NUMBER_FORMAT_EUR
                elif column_name in current_percentage_columns:
                    target_format = PERCENTAGE_NUMBER_FORMAT
                else:
                    continue

                column_letter = _column_index_to_excel_letters(column_index)
                for row_number in range(2, number_of_rows + 2):
                    target_cell = current_worksheet[f"{column_letter}{row_number}"]
                    target_cell.number_format = target_format

            if current_col_width is not None:
                for column_index in range(1, number_of_columns + 1):
                    column_letter = _column_index_to_excel_letters(column_index)
                    current_worksheet.column_dimensions[column_letter].width = current_col_width

    print(f"File saved successfully: {output_file_path}")
    for configuration in sheet_configurations:
        print(f'\t- {configuration.sheet_name:<10} -> {len(configuration.df):>6} rows')

